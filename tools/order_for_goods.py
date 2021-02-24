import openpyxl, sys, pprint, logging
from openpyxl.utils import get_column_letter, column_index_from_string

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s : %(message)s')

db_path = 'C:/Users/asus/Desktop/分仓系统需求-20210122.xlsx'
db_save = 'C:/Users/asus/Desktop/分仓系统需求-20210122_bak.xlsx'
goods_sheet = '亚马逊订单原文件'
product_sheet = '平台SKU及产品SKU对应表'
zone_sheets = ['孟菲斯仓邮编分区表', '洛杉矶仓邮编分区表']
price_sheets = ['孟菲斯仓分区价格表', '洛杉矶仓分区价格表']

goods_sku = []
goods_code = []
product_sku = {}
zone = {}
price = {}

wb = openpyxl.load_workbook(db_path)

def collect_goods(goods_sheet):
    sheet = wb.get_sheet_by_name(goods_sheet)
    if sheet == None:
        return
    for row in range(2, sheet.max_row + 1):
        goods_sku.append(sheet['K' + str(row)].value)
        goods_code.append(sheet['W' + str(row)].value)

def collect_product(product_sheet):
    sheet = wb.get_sheet_by_name(product_sheet)
    if sheet == None:
        return
    for row in range(2, sheet.max_row + 1):
        sku = sheet['E' + str(row)].value
        product_sku[sku] = 0
        count = int(sheet['F' + str(row)].value)
        if count > 0:
            for i in range(0, count):
                weight = sheet[get_column_letter(3 * i + 9) + str(row)].value
                if weight == None:
                    continue
                weight = int(weight)
                product_sku[sku] += weight

def collect_zone(zone_sheets):
    for sheet_name in zone_sheets:
        sheet = wb.get_sheet_by_name(sheet_name)
        if sheet == None:
            continue
        zone.setdefault(sheet_name, {})
        for row in range(2, sheet.max_row + 1):
            dest_zip = str(sheet['A' + str(row)].value)
            zone_num = str(sheet['B' + str(row)].value)
            postal_code = dest_zip.split('-')
            if len(postal_code) == 2:
                for a_code in range(int(postal_code[0]), int(postal_code[1]) + 1):
                    zone[sheet_name][a_code] = zone_num
            else:
                zone[sheet_name][int(dest_zip)] = zone_num

def collect_price(price_sheets):
    for sheet_name in price_sheets:
        sheet = wb.get_sheet_by_name(sheet_name)
        if sheet == None:
            continue
        price.setdefault(sheet_name, {})
        for col in range(2, sheet.max_column + 1):
            col_idx = get_column_letter(col)
            zone_num = sheet[col_idx + str(1)].value
            price[sheet_name].setdefault(zone_num, {})
            for row in range(2, sheet.max_row + 1):
                price[sheet_name][zone_num][sheet['A' + str(row)].value] = sheet[col_idx + str(row)].value

def save_warehouse(db_save, warehouse):
    sheet = wb.get_sheet_by_name(goods_sheet)
    if sheet == None:
        return
    if len(warehouse) != sheet.max_row - 1:
        logging.warning('len(warehouse) != sheet.max_row - 1')
        return
    for row in range(2, sheet.max_row + 1):
        sheet['AF' + str(row)] = warehouse[row - 2]
    wb.save(db_save)

if __name__ == '__main__':
    collect_goods(goods_sheet)
    collect_product(product_sheet)
    collect_zone(zone_sheets)
    collect_price(price_sheets)
    if len(goods_sku) != len(goods_code):
        logging.warning('len(goods_sku) != len(goods_code)')
    elif len(zone_sheets) != len(price_sheets):
        logging.warning('len(zone_sheets) != len(price_sheets)')
    else:
        weight_key = 0
        destzip_key = -1
        current_idx = -1
        zone_ids = []
        warehouse = []
        # 遍历以处理所有的订单
        for sku in goods_sku:
            current_idx += 1
            if sku in product_sku.keys():
                weight_key = product_sku[sku]
                destzip_key = int(str(goods_code[current_idx])[0:3])
            else:
                warehouse.append('')
                continue
            for zone_name in zone_sheets:
                zone_id = zone[zone_name].get(destzip_key, 'NA')
                if zone_id == 'NA':
                    zone_id = -1
                else:
                    zone_id = int(zone_id)
                zone_ids.append(zone_id)
            current_idx = -1
            warehouse_of_minprice = ''
            price_minval = sys.float_info.max
            for price_name in price_sheets:
                current_idx += 1
                zone_id = zone_ids[current_idx]
                if zone_id != -1 and zone_id in price[price_name].keys() and weight_key in price[price_name][zone_id].keys():
                    cur_price = float(price[price_name][zone_id][weight_key])
                    if cur_price < price_minval:
                        price_minval = cur_price
                        warehouse_of_minprice = price_name
            warehouse.append(warehouse_of_minprice)
        save_warehouse(db_save, warehouse)
